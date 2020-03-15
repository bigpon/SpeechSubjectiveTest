import sys
import os
import random
import time
import platform
import winsound
from playsound import playsound
from openpyxl import Workbook
from openpyxl import load_workbook

def playspeech(filename):
    if os.path.exists(filename):
        if platform.system() == 'Windows':
            winsound.PlaySound(filename, winsound.SND_ALIAS)
        else:
            playsound(filename)
    else:
        print("%s doesn't exists!!"%filename)
        sys.exit(0)

def introduction(test_type, text, examples):
    """TEST INTRODUCTION
    Args:
        test_type (str): the type of the test
    """
    if test_type == 'PK':
        info = "'Each PK' test will play two files," + \
               " and please select the answer" + \
               " with %s %s.\n" % (text['PK']['Ans'], text['PK']['C'])
        print(info) 
    elif test_type == 'SIM':
        info = "Each 'SIM' test will play two files" + \
               " and please select the answer" + \
               " correspoinding to the question.\n"
        print(info) 
    elif test_type == 'MOS':
        info = "Each 'MOS' test will play one file," + \
               " and please select the score" + \
               " correspoinding to the question."
        print(info)
        print("There are several examples with score 5\n")
        input("Press any key to play the examples!\n")
        for i, file in enumerate(examples['MOS']):
            print("example%d" % i)
            playspeech(file)
        print('\n')
    elif test_type == 'XAB':
        info = "Each 'XAB' test will play 'reference', 'A'" + \
               " and 'B' files, please selct the %s" % text['XAB']['C']+ \
               " of 'A' or 'B' is %s to that of the reference." % text['XAB']['Ans']
        print(info)
        print("There is an example\n")
        input("Press any key to play the example!\n")
        print("example reference")
        playspeech(examples['XAB']['ref'])
        print("postive example (correct answer)")
        playspeech(examples['XAB']['pos'])
        print("negative example (wrong answer)")
        playspeech(examples['XAB']['neg'])
        print('\n')
    else:
        raise ValueError(
            "test type %s is not in (MOS, SIM, PK, XAB)!"%(test_type))

# XAB test
def main_XAB(eval_dict, r_idx, text):
    """MAIN OF XAB TEST
        Args:
            eval_dict (dict): evaluation dict
            r_idx (int): index of unfinished files
            text (dict): dict of text information
    """
    # MethodA(expected answer) v.s. MethodB with the reference MethodX
    fileX = eval_dict['FileX']
    answer = random.randint(0, 1)
    #print(answer)
    if answer:
        # expected answer is fileA
        fileA = eval_dict['FileA']
        fileB = eval_dict['FileB']
    else:
        # expected answer is fileB
        fileA = eval_dict['FileB']
        fileB = eval_dict['FileA']
    #print(fileX)
    #print(fileA)
    #print(fileB)
    print("*"+("[No. %4d]-0 reference wav file" % (r_idx)).center(20)+"*")
    playspeech(fileX)
    time.sleep(0.3)
    print("*"+("[No. %4d]-1 'A' wav file" % (r_idx)).center(20)+"*")
    playspeech(fileA)
    time.sleep(0.3)
    print("*"+("[No. %4d]-2 'B' wav file" % (r_idx)).center(20)+"*")
    playspeech(fileB)
    ans_flag = True
    Criteria = text['XAB']['C']
    Ans = text['XAB']['Ans']
    Q = text['XAB']['Q'].replace('Ans', Ans).replace('Criteria', Criteria)
    while ans_flag:
        print("\n%s" % Q)
        ans = input(("%-6s[1].%-22s[2].%-22s\n" %
                     ("Enter ", " 'A' is %s" % Ans, " 'B' is %s" % Ans)) + \
                    ("%-6s[6].%-22s[7].%-22s\n" %
                     ("", " replay 'A' wav", " replay 'B' wav")) + \
                    ("%-6s[8].%-22s[9].%-22s\n" %
                     ("", " replay reference wav", " replay all files")) + \
                    ("%-6s[q].%-22s\n: " %
                     ("", " quit evaluation")))
                    
        if ans == "1":
            eval_dict['Score'] = answer
            eval_dict['Finished'] = True
            ans_flag = False
        elif ans == "2":
            eval_dict['Score'] = (1 - answer)
            eval_dict['Finished'] = True
            ans_flag = False
        elif ans == "6":
            print("'A' wav")
            playspeech(fileA)
        elif ans == "7":
            print("'B' wav")
            playspeech(fileB)
        elif ans == "8":
            print("reference wav")
            playspeech(fileX)
        elif ans == "9":
            print("reference wav")
            playspeech(fileX)
            time.sleep(0.3)
            print("'A' wav")
            playspeech(fileA)
            time.sleep(0.3)
            print("'B' wav")
            playspeech(fileB)
        elif ans == "q":
            return ans
        else:
            print("ans should be 1,2,6,7,8,9 or q. %s is not included!!" % ans)
    return ans

# preference test
def main_PK(eval_dict, r_idx, text):
    """MAIN OF PK TEST
        Args:
            eval_dict (dict): evaluation dict
            r_idx (int): index of unfinished files
            text (dict): dict of text information
    """
    # MethodA(expected answer) v.s. MethodB
    answer = random.randint(0, 1)
    #print(answer)
    if answer:
        # expected answer is fileA
        fileA = eval_dict['FileA']
        fileB = eval_dict['FileB']
    else:
        # expected answer is fileB
        fileA = eval_dict['FileB']
        fileB = eval_dict['FileA']
    #print(fileA)
    #print(fileB)
    print("*"+("[No. %4d]-1 'A' wav file" % (r_idx)).center(20)+"*")
    playspeech(fileA)
    time.sleep(0.3)
    print("*"+("[No. %4d]-2 'B' wav file" % (r_idx)).center(20)+"*")
    playspeech(fileB)
    ans_flag = True
    Criteria = text['PK']['C']
    Ans = text['PK']['Ans']
    Q = text['PK']['Q'].replace('Ans', Ans).replace('Criteria', Criteria)
    while ans_flag:
        print("\n%s" % Q)
        ans = input(("%-6s[1].%-22s[2].%-22s\n" %
                     ("Enter ", " 'A' is %s" % Ans, " 'B' is %s" % Ans)) + \
                    ("%-6s[7].%-22s[8].%-22s\n" %
                     ("", " replay 'A' wav", " replay 'B' wav")) + \
                    ("%-6s[9].%-22s[q].%-22s\n: " %
                     ("", " replay both files", " quit evaluation")))
        if ans == "1":
            eval_dict['Score'] = answer
            eval_dict['Finished'] = True
            ans_flag = False
        elif ans == "2":
            eval_dict['Score'] = (1 - answer)
            eval_dict['Finished'] = True
            ans_flag = False
        elif ans == "7":
            print("'A' wav")
            playspeech(fileA)
        elif ans == "8":
            print("'B' wav")
            playspeech(fileB)
        elif ans == "9":
            print("'A' wav")
            playspeech(fileA)
            time.sleep(0.3)
            print("'B' wav")
            playspeech(fileB)
        elif ans == "q":
            return ans
        else:
            print("ans should be 1,2,7,8,9 or q. %s is not included!!" % ans)
    return ans

# MOS test
def main_MOS(eval_dict, r_idx, text): 
    """MAIN OF MOS TEST
        Args:
            eval_dict (dict): evaluation dict
            r_idx (int): index of unfinished files
            text (dict): dict of text information
    """ 
    # PLAY FILES
    file = eval_dict['File']
    #print(file)
    print("*"+("[No. %4d] wav file" % (r_idx)).center(20)+"*")
    playspeech(file)
    # INPUT ANSWER
    ans_flag = True
    scores = ["1", "2", "3", "4", "5"]
    Criteria = text['MOS']['C']
    Q = text['MOS']['Q'].replace('Criteria', Criteria)
    while ans_flag:
        print("\n%s" % Q)
        ans = input(("%-6s[1].%-22s[2].%-22s\n" %
                     ("Enter ", " Bad", " Poor")) + \
                    ("%-6s[3].%-22s[4].%-22s\n" %
                     ("", " Fair", " Good")) + \
                    ("%-6s[5].%-22s[7].%-22s\n" %
                     ("", " Excellent", " replay wav")) + \
                    ("%-6s[q].%-22s\n: " %
                     ("", " quit evaluation")))
        if ans in scores:
            eval_dict['Score'] = float(ans)
            eval_dict['Finished'] = True
            ans_flag = False
        elif ans == "7":
            print("replay wav")
            playspeech(file)
        elif ans == "q":
            return ans
        else:
            print("ans should be 1,2,3,4,5,7 or q. %s is not included!!" % ans)
    return ans

# similarity test
def main_SIM(eval_dict, r_idx, text):
    """MAIN OF SIM TEST
        Args:
            eval_dict (dict): evaluation dict
            r_idx (int): index of unfinished files
            text (dict): dict of text information
    """
    # PLAY FILES
    if random.randint(0, 1):
        file1 = eval_dict['File']
        file2 = eval_dict['File_ans']
    else:
        file1 = eval_dict['File_ans']
        file2 = eval_dict['File']
    #print(file1)
    #print(file2)
    print("*"+("[No. %4d]-1 wav file" % (r_idx)).center(20)+"*")
    playspeech(file1)
    time.sleep(0.1)
    print("*"+("[No. %4d]-2 wav file" % (r_idx)).center(20)+"*")
    playspeech(file2)
    # INPUT ANSWER
    ans_flag = True
    scores = ["1", "2", "3", "4"]
    Q = text['SIM']['Q']
    posAns = text['SIM']['posAns']
    negAns = text['SIM']['negAns']
    while ans_flag:
        print("\n%s" % Q)
        ans = input(("%-6s[1].%-22s[2].%-22s\n" %
                     ("Enter ", " definitely %s" % posAns, " maybe %s" % posAns)) + \
                    ("%-6s[3].%-22s[4].%-22s\n" %
                     ("", " maybe %s" % negAns, " definitely %s" % negAns)) + \
                    ("%-6s[7].%-22s[8].%-22s\n" %
                     ("", " replay 1st wav", " replay 2nd wav")) + \
                    ("%-6s[9].%-22s[q].%-22s\n: " %
                     ("", " replay both files", " quit evaluation")))
        if ans in scores:
            eval_dict['Score'] = float(ans)
            eval_dict['Finished'] = True
            ans_flag = False
        elif ans == "7":
            print("1st wav")
            playspeech(file1)
        elif ans == "8":
            print("2nd wav")
            playspeech(file2)
        elif ans == "9":
            print("1st wav")
            playspeech(file1)
            time.sleep(0.1)
            print("2nd wav")
            playspeech(file2)
        elif ans == "q":
            return ans
        else:
            print("ans should be 1,2,3,4,7,8,9 or q. %s is not included!!" % ans)
    return ans

# XLSX OUTPUT TEMPLATE 
class xlsx_output(object):
    def __init__(self, username, filename, sheetnames, testsystems):
        self.u_name = username
        self.fxlsx = filename
        self.t_sheets = sheetnames
        self.t_systems = testsystems
        # column index of xslx (alphabet list = ['A', 'B', ..., 'Z'])
        self.alphabet = []
        for i in range(26):
            self.alphabet.append(chr(ord('A')+i))

    def _add_username(self, sheet, c_row, t_name):
        # add user in the new row
        sheet.cell(row=c_row, column=1).value = t_name
        sheet.cell(row=(c_row+1), column=1).value = 'AVG.'

    def _add_data(self, sheet, c_row, col_chr, t_score):
        # add new scores
        sheet['%s%d' % (col_chr, c_row)].value = t_score
        start_idx = "%s%d" % (col_chr, 2)
        end_idx = "%s%d" % (col_chr, c_row)
        # update the average scores
        sheet['%s%d' % (col_chr, c_row+1)].value = \
            "=AVERAGE(%s:%s)" % (start_idx, end_idx)

# XAB, PREFERENCE, MOS XLSX OUTPUT CLASS
class xlsx_SCORE(xlsx_output):
    """
        Class of preference, MOS, and XAB tests
    """
    def __init__(self, username, filename, sheetnames, testsystems):
        super().__init__(username, filename, sheetnames, testsystems)
        if not os.path.exists(self.fxlsx):
            self._initial_xlsx()

    def output_xlsx(self, t_sheet, t_dict):
        """OUTPUT SUBJECTIVE RESULTS INTO A SHEET OF EXCEL FILE
        Args:
            t_sheet (str): the subset name of the subjective results
                [summary, xgender, sgender, F-F, F-M, M-F, M-M]
            t_dict (list of dict): the result dicts 
        """
        if len(t_dict) == 0:
            print("%s is empty!!" % t_sheet)
            return
        wb = load_workbook(self.fxlsx) # load workspace of the excel file
        sheet = wb['%s' % t_sheet] # load sheet
        c_row = sheet.max_row # get latest row index
        # add new user
        self._add_username(sheet, c_row, self.u_name)
        # parse results
        t_score = self._score(t_dict)
        # update sheet
        for i in range(len(t_score)):
            self._add_data(sheet, c_row, self.alphabet[i+1], t_score[i])
        wb.save(self.fxlsx)

    def _initial_xlsx(self):
        wb = Workbook()
        first = True
        for s_name in self.t_sheets:
            if first:
                sheet = wb.active
                sheet.title = s_name
                first = False
            else:
                wb.create_sheet(title=s_name)
                sheet = wb['%s' % s_name]
            sheet['A1'].value = 'USER'
            sheet['A2'].value = 'AVG.'
            for i in range(len(self.t_systems)):
                sheet.cell(row=1, column=(2+i)).value = self.t_systems[i]
        wb.save(self.fxlsx)

    def _score(self, t_dict):
        t_score = []
        for t_method in self.t_systems:
            score = 0.0
            t_num = 0
            f_t_dict = filter(lambda item: item['method'] == t_method, t_dict)
            for t_file in f_t_dict:
                score += t_file['Score']
                t_num+=1
            if t_num==0:
                t_score.append(-1.0)
            else:
                t_score.append(score/t_num)
        return t_score

# SIMILARITY XLSX OUTPUT CLASS
class xlsx_SIM(xlsx_output):
    """
        Class of similarity test
    """
    def __init__(self, username, filename, sheetnames, testsystems, text):
        super().__init__(username, filename, sheetnames, testsystems)
        self.posAns_sure = '%s' % text['SIM']['posAns']
        self.posAns_weak = 'maybe_%s' % text['SIM']['posAns']
        self.negAns_weak = 'maybe_%s' % text['SIM']['negAns']
        self.negAns_sure = '%s' % text['SIM']['negAns']
        if not os.path.exists(self.fxlsx):
            self._initial_xlsx()
    
    def output_xlsx(self, t_sheet, t_dict):
        """OUTPUT SUBJECTIVE RESULTS INTO A SHEET OF EXCEL FILE
        Args:
            t_sheet (str): the subset name of the subjective results
                [summary, xgender, sgender, F-F, F-M, M-F, M-M]
            t_dict (list of dict): the result dicts 
        """
        if len(t_dict) == 0:
            print("%s is empty!!" % t_sheet)
            return
        wb = load_workbook(self.fxlsx) # load workspace of the excel file
        sheet = wb['%s' % t_sheet] # load sheet
        c_row = sheet.max_row # get latest row index
        # add new user
        self._add_username(sheet, c_row, self.u_name)
        # parse results
        t_score = self._score(t_dict)
        # update sheet
        for i in range(len(t_score)):
            self._add_data(sheet, c_row, self.alphabet[i*4+1], t_score[i][self.posAns_sure])
            self._add_data(sheet, c_row, self.alphabet[i*4+2], t_score[i][self.posAns_weak])
            self._add_data(sheet, c_row, self.alphabet[i*4+3], t_score[i][self.negAns_weak])
            self._add_data(sheet, c_row, self.alphabet[i*4+4], t_score[i][self.negAns_sure])
        #self._final_result(sheet)
        wb.save(self.fxlsx)

    def _final_result(self, sheet):
        c_row = sheet.max_row + 1
        sheet.cell(row=c_row, column=1).value = 'Result'
        m_idx = range(1, len(self.t_systems)*4+1, 4)
        for i in range(len(self.t_systems)):
            # SAME SPEAKER
            merge_range = "%s%d:%s%d" % (
                chr(ord('A')+m_idx[i]), c_row, chr(ord('B')+m_idx[i]), c_row)
            sheet.merge_cells(merge_range)
            sum_range = "%s%d:%s%d" % (
                chr(ord('A')+m_idx[i]), c_row-1, chr(ord('B')+m_idx[i]), c_row-1)
            sheet.cell(row=c_row, column=(2+i*4)).value = "=SUM(%s)" % (sum_range)
            # DIFFERENT SPEAKER
            merge_range = "%s%d:%s%d" % (
                chr(ord('C')+m_idx[i]), c_row, chr(ord('D')+m_idx[i]), c_row)
            sheet.merge_cells(merge_range)
            sum_range = "%s%d:%s%d" % (
                chr(ord('C')+m_idx[i]), c_row-1, chr(ord('D')+m_idx[i]), c_row-1)
            sheet.cell(row=c_row, column=(4+i*4)).value = "=SUM(%s)" % (sum_range)
            
    def _initial_xlsx(self):
        wb = Workbook()
        first = True
        for s_name in self.t_sheets:
            if first:
                sheet = wb.active
                sheet.title = s_name
                first = False
            else:
                wb.create_sheet(title=s_name)
                sheet = wb['%s' % s_name]
            sheet['A1'].value = 'USER'
            sheet['A2'].value = ''
            sheet['A3'].value = 'AVG.'
            m_idx = range(1, len(self.t_systems)*4+1, 4)
            for i in range(len(self.t_systems)):
                merge_range = "%s%d:%s%d" % (chr(ord('A')+m_idx[i]), 1, chr(ord('D')+m_idx[i]), 1)
                sheet.merge_cells(merge_range)
                sheet.cell(row=1, column=(2+i*4)).value = self.t_systems[i]
                sheet.cell(row=2, column=(2+i*4)).value = self.posAns_sure
                sheet.cell(row=2, column=(3+i*4)).value = self.posAns_weak
                sheet.cell(row=2, column=(4+i*4)).value = self.negAns_weak
                sheet.cell(row=2, column=(5+i*4)).value = self.negAns_sure
        wb.save(self.fxlsx)

    def _score(self, t_dict):
        t_score = []
        for t_method in self.t_systems:
            score = {self.posAns_sure: 0.0, self.posAns_weak: 0.0, 
                     self.negAns_sure: 0.0, self.negAns_weak: 0.0}
            f_t_dict = filter(lambda item: item['method'] == t_method, t_dict)
            t_num = 0
            for t_file in f_t_dict:
                if t_file['Score'] == 1.0:
                    ans = self.posAns_sure
                elif t_file['Score'] == 2.0:
                    ans = self.posAns_weak
                elif t_file['Score'] == 3.0:
                    ans = self.negAns_weak
                elif t_file['Score'] == 4.0:
                    ans = self.negAns_sure
                else:
                    raise ValueError("%s method score %.2f is out of range" % (
                        t_method, t_file['Score']))
                score[ans] += 1.0
                t_num += 1
            if t_num == 0:
                t_score.append(score)
            else:
                score = {key: value / t_num for key, value in score.items()}
                t_score.append(score)
        return t_score