#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright 2020 Wu Yi-Chiao (Nagoya University)
# Apache 2.0  (http://www.apache.org/licenses/LICENSE-2.0)

import os
import sys
import numpy as np
import yaml
import scipy.stats
import copy
from openpyxl import Workbook
from openpyxl import load_workbook

# STATISTICAL RESULTS CLASS
class StatisticResult(object):
    """
        Class to output statistics of evaluation results
    """
    def __init__(self, c_xlsx, sheetname):
        self.c_xlsx = c_xlsx
        self.t_sheet = sheetname
        self.mean = []
        self.std = []
        self.ci = []
        self.clow = []
        self.cup = []
        self.data_length = 0

    def push_result(self, scores, confidence=0.95):
        """PUSH SUBJECTIVE RESULTS OF A USER, PAIR, OR SYSTEM TO THE QUEUE OF OVERALL RESULTS
        Args:
            scores (list): score list
            confidence (float): the level of confidence (0~1.0) (default: 0.95 (95%)) 
        """
        self.data_length = len(scores)
        if self.data_length == 0:
            #print('(push) data_length of %s is 0!' % self.t_sheet)
            return 0
        mean, std, ci, clow, cup = self._mean_confidence_interval(scores, confidence)
        self.mean.append(mean)
        self.std.append(std)
        self.ci.append(ci)
        self.clow.append(clow)
        self.cup.append(cup)

    def output_result(self):
        """OUTPUT STATISTICS OF ALL EVALUATION RESULTS INTO A EXCEL FILE
            mean: average of results
            std: standard deviation of results 
            ci: confidence interval
            clow: lowerbound of the confidence interval
            cup: upperbound of the confidence interval
        """
        if self.data_length == 0:
            return 0
        self.c_xlsx.output_xlsx('mean', self.t_sheet, self.mean)
        self.c_xlsx.output_xlsx('std',  self.t_sheet, self.std)
        self.c_xlsx.output_xlsx('ci',   self.t_sheet, self.ci)
        self.c_xlsx.output_xlsx('clow', self.t_sheet, self.clow)
        self.c_xlsx.output_xlsx('cup',  self.t_sheet, self.cup)

    def _mean_confidence_interval(self, data, confidence=0.95):
        a = 1.0 * np.array(data)
        #n = len(a)
        mean = np.mean(a)
        std = np.std(a)
        std_err = scipy.stats.sem(a)
        ci = std_err * scipy.stats.norm.ppf((1 + confidence) / 2.) # confidence interval
        return mean, std, ci, mean-ci, mean+ci, 

# XLSX FINAL OUTPUT TEMPLATE 
class xlsx_foutput(object):
    def __init__(self, filename, sheetnames, testsystems):
        self.fxlsx = filename
        self.t_sheets = sheetnames
        self.t_systems = testsystems
        # column index of xslx (alphabet list = ['A', 'B', ..., 'Z'])
        self.alphabet = []
        for i in range(26):
            self.alphabet.append(chr(ord('A')+i))

    def _add_username(self, sheet, c_row, t_name):
        # add user in the new row
        sheet.cell(row=c_row, column=1).value = t_name

    def _add_data(self, sheet, c_row, col_chr, t_score):
        # add new scores
        sheet['%s%d' % (col_chr, c_row)].value = t_score

# XAB, PREFERENCE, MOS XLSX FINAL OUTPUT CLASS
class xlsx_fSCORE(xlsx_foutput):
    """
        Class to output evaluation results
    """
    def __init__(self, filename, sheetnames, testsystems, c_text):
        super().__init__(filename, sheetnames, testsystems)
        if os.path.exists(self.fxlsx):
            print("overwrite %s" % self.fxlsx)
        self._initial_xlsx()
        # initialize score list dictionary of each test system
        system_dict = {}
        for t_system in self.t_systems:
            system_dict = {**system_dict, t_system:[]}  
        # initialize score list dictionary of each test pair and test system
        self.user_score={} # user-based score
        self.utt_score={} # utterance-based score
        for t_pair in self.t_sheets:
            self.user_score = {**self.user_score, t_pair:copy.deepcopy(system_dict)}
            self.utt_score = {**self.utt_score, t_pair:copy.deepcopy(system_dict)}
        # load config text
        self.c_text = c_text

    def output_xlsx(self, t_name, t_sheet, t_score):
        """OUTPUT SUBJECTIVE RESULTS INTO A SHEET OF EXCEL FILE
        Args:
            t_name (str): the user name
            t_sheet (str): the subset name of the subjective results
                [summary, xgender, sgender, F-F, F-M, M-F, M-M]
            t_score (list of score): the list of scores of the subset
        """
        if len(t_score) == 0:
            print("%-10s - %-6s is empty!!" % (t_name, t_sheet))
            return
        wb = load_workbook(self.fxlsx) # load workspace of the excel file
        sheet = wb['%s' % t_sheet] # load sheet
        c_row = sheet.max_row # get latest row index
        c_row += 1
        # add new user
        self._add_username(sheet, c_row, t_name)
        # update sheet
        for i, score in enumerate(t_score):
            self._add_data(sheet, c_row, self.alphabet[i+1], score)
        wb.save(self.fxlsx)
    
    def output_result(self, t_name, t_sheet, t_dict):
        """OUTPUT SUBJECTIVE RESULTS
        Args:
            t_name (str): the user name
            t_sheet (str): the subset name of the subjective results
                [summary, xgender, sgender, F-F, F-M, M-F, M-M]
            t_dict (list of dict): the result list of dicts of the subset
        """
        if len(t_dict) == 0:
            print("%-10s - %-6s is empty!!" % (t_name, t_sheet))
            return
        else:
            # parse results
            t_score = self._score(t_sheet, t_dict)
            self.output_xlsx(t_name, t_sheet, t_score)

    def _initial_xlsx(self):
        wb = Workbook()
        first = True
        for t_pair in self.t_sheets:
            if first:
                sheet = wb.active
                sheet.title = t_pair
                first = False
            else:
                wb.create_sheet(title=t_pair)
                sheet = wb['%s' % t_pair]
            sheet['A1'].value = 'USER'
            for i in range(len(self.t_systems)):
                sheet.cell(row=1, column=(2+i)).value = self.t_systems[i]
        wb.save(self.fxlsx)
    
    def _score(self, t_pair, t_dict):
        t_score = []
        for t_system in self.t_systems:
            score = 0.0
            t_num = 0
            f_t_dict = filter(lambda item: item[self.c_text['system']] == t_system, t_dict)
            for t_file in f_t_dict:
                u_score = t_file[self.c_text['r_score']] # score of the utterance
                self.utt_score[t_pair][t_system] += [u_score]
                score += u_score
                t_num+=1
            if t_num==0:
                score = -1.0
                self.utt_score[t_pair][t_system] += [-1.0]
            else:
                score /= t_num
            t_score.append(score)
            # update score of each system under this test pair
            self.user_score[t_pair][t_system] += [score]
        return t_score

